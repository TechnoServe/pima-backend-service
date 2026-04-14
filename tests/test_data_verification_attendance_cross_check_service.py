from __future__ import annotations

import io
import unittest
from datetime import date
from uuid import UUID, uuid4

from openpyxl import load_workbook

from app.domains.data_verification.service import DataVerificationService
from app.shared.api_errors import DomainError, NotFoundError, ValidationError


class FakeDataVerificationRepo:
    def __init__(self, project_id: UUID, training_group_id: UUID, other_group_id: UUID):
        self.project_id = project_id
        self.training_group_id = training_group_id
        self.other_group_id = other_group_id
        self.last_latest_call: dict | None = None

        f1 = uuid4()
        f2 = uuid4()
        f3 = uuid4()

        self.latest_rows = [
            {
                "id": uuid4(),
                "farmer_id": f1,
                "observation_id": None,
                "farm_visit_id": uuid4(),
                "training_session_id": uuid4(),
                "date_completed": date(2026, 2, 10),
                "attended_trainings": True,
                "number_of_trainings_attended": 1,
                "attended_last_months_training": "Yes",
                "check_type": "Farm Visit",
                "first_name": "Alice",
                "middle_name": "A",
                "last_name": "Farmer",
                "tns_id": "TNS-100",
                "training_group_id": training_group_id,
                "training_group_name": "Group One",
                "training_module_id": uuid4(),
                "training_module_name": "Module 2",
                "training_module_number": 2,
            },
            {
                "id": uuid4(),
                "farmer_id": f2,
                "observation_id": uuid4(),
                "farm_visit_id": None,
                "training_session_id": uuid4(),
                "date_completed": date(2026, 2, 11),
                "attended_trainings": True,
                "number_of_trainings_attended": 1,
                "attended_last_months_training": "Yes",
                "check_type": "Training Observation",
                "first_name": "Bob",
                "middle_name": None,
                "last_name": "River",
                "tns_id": "TNS-200",
                "training_group_id": training_group_id,
                "training_group_name": "Group One",
                "training_module_id": uuid4(),
                "training_module_name": "Module 2",
                "training_module_number": 2,
            },
            {
                "id": uuid4(),
                "farmer_id": f3,
                "observation_id": None,
                "farm_visit_id": None,
                "training_session_id": uuid4(),
                "date_completed": date(2026, 2, 12),
                "attended_trainings": True,
                "number_of_trainings_attended": 2,
                "attended_last_months_training": "No training was offered",
                "check_type": "Regular",
                "first_name": "Chris",
                "middle_name": None,
                "last_name": "Lake",
                "tns_id": "TNS-300",
                "training_group_id": training_group_id,
                "training_group_name": "Group One",
                "training_module_id": uuid4(),
                "training_module_name": "Module 1",
                "training_module_number": 1,
            },
        ]

        self.attendance = {
            f1: [
                {
                    "attendance_id": uuid4(),
                    "training_session_id": uuid4(),
                    "training_date": date(2026, 1, 10),
                    "module_id": uuid4(),
                    "module_name": "Module 1",
                    "module_number": 1,
                    "current_previous": "current",
                    "status": "Present",
                },
                {
                    "attendance_id": uuid4(),
                    "training_session_id": uuid4(),
                    "training_date": date(2026, 2, 10),
                    "module_id": uuid4(),
                    "module_name": "Module 2",
                    "module_number": 2,
                    "current_previous": "current",
                    "status": "Present",
                },
            ],
            f2: [
                {
                    "attendance_id": uuid4(),
                    "training_session_id": uuid4(),
                    "training_date": date(2026, 2, 8),
                    "module_id": uuid4(),
                    "module_name": "Module 2",
                    "module_number": 2,
                    "current_previous": "current",
                    "status": "Present",
                }
            ],
            f3: [
                {
                    "attendance_id": uuid4(),
                    "training_session_id": uuid4(),
                    "training_date": date(2026, 2, 1),
                    "module_id": uuid4(),
                    "module_name": "Module 1",
                    "module_number": 1,
                    "current_previous": "current",
                    "status": "Present",
                },
                {
                    "attendance_id": uuid4(),
                    "training_session_id": uuid4(),
                    "training_date": date(2026, 2, 2),
                    "module_id": uuid4(),
                    "module_name": "Module 1",
                    "module_number": 1,
                    "current_previous": "current",
                    "status": "Present",
                },
            ],
        }

    async def project_exists(self, project_id: UUID) -> bool:
        return project_id == self.project_id

    async def training_group_belongs_to_project(self, project_id: UUID, training_group_id: UUID) -> bool:
        return project_id == self.project_id and training_group_id == self.training_group_id

    async def get_latest_checks_for_attendance_cross_check(self, *, project_id: UUID, search: str | None, training_group_id: UUID | None, verification_source: str):
        self.last_latest_call = {
            "project_id": project_id,
            "search": search,
            "training_group_id": training_group_id,
            "verification_source": verification_source,
        }
        rows = list(self.latest_rows)
        if training_group_id is not None:
            rows = [r for r in rows if r["training_group_id"] == training_group_id]
        if search:
            q = search.lower()
            rows = [
                r
                for r in rows
                if q in (r.get("first_name") or "").lower()
                or q in (r.get("middle_name") or "").lower()
                or q in (r.get("last_name") or "").lower()
                or q in (r.get("tns_id") or "").lower()
            ]
        if verification_source == "farm_visit":
            rows = [r for r in rows if r["farm_visit_id"] is not None]
        elif verification_source == "training_observation":
            rows = [r for r in rows if r["observation_id"] is not None and r["farm_visit_id"] is None]
        elif verification_source == "none":
            rows = [r for r in rows if r["farm_visit_id"] is None and r["observation_id"] is None]
        return rows

    async def get_attendance_evidence_for_farmers(self, *, project_id: UUID, farmer_ids: list[UUID]):
        return {farmer_id: self.attendance.get(farmer_id, []) for farmer_id in farmer_ids}


class AttendanceCrossCheckServiceTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self):
        self.project_id = uuid4()
        self.training_group_id = uuid4()
        self.other_group_id = uuid4()
        self.repo = FakeDataVerificationRepo(self.project_id, self.training_group_id, self.other_group_id)
        self.service = DataVerificationService(db=None)
        self.service.repo = self.repo

    async def test_successful_list_response(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search=None,
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertEqual(response.totals.total, 3)
        self.assertEqual(response.totals.matches, 2)
        self.assertEqual(response.totals.mismatches, 1)

    async def test_empty_result(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="not-found",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertEqual(response.items, [])
        self.assertEqual(response.totals.total, 0)

    async def test_invalid_verification_source(self):
        with self.assertRaises(ValidationError):
            await self.service.list_attendance_cross_check(
                project_id=self.project_id,
                page=1,
                page_size=10,
                search=None,
                training_group_id=None,
                verification_source="bad",
                only_mismatches=False,
            )

    async def test_missing_project(self):
        with self.assertRaises(NotFoundError):
            await self.service.list_attendance_cross_check(
                project_id=uuid4(),
                page=1,
                page_size=10,
                search=None,
                training_group_id=None,
                verification_source="all",
                only_mismatches=False,
            )

    async def test_training_group_outside_project(self):
        with self.assertRaises(DomainError) as exc:
            await self.service.list_attendance_cross_check(
                project_id=self.project_id,
                page=1,
                page_size=10,
                search=None,
                training_group_id=self.other_group_id,
                verification_source="all",
                only_mismatches=False,
            )
        self.assertEqual(exc.exception.status_code, 422)

    async def test_search_by_tns_id(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="tns-200",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].tns_id, "TNS-200")

    async def test_search_by_farmer_name(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="alice",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].first_name, "Alice")

    async def test_filter_by_farm_visit(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search=None,
            training_group_id=None,
            verification_source="farm_visit",
            only_mismatches=False,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].comparison_rule, "farm_visit")

    async def test_filter_by_training_observation(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search=None,
            training_group_id=None,
            verification_source="training_observation",
            only_mismatches=False,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].comparison_rule, "training_observation")

    async def test_filter_by_none(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search=None,
            training_group_id=None,
            verification_source="none",
            only_mismatches=False,
        )
        self.assertEqual(len(response.items), 1)
        self.assertEqual(response.items[0].comparison_rule, "full")

    async def test_farm_visit_comparison_rule(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Alice",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        item = response.items[0]
        self.assertIsNone(item.matches.count_equal)
        self.assertTrue(item.matches.any_equal)

    async def test_training_observation_comparison_rule(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Bob",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        item = response.items[0]
        self.assertIsNone(item.matches.any_equal)
        self.assertFalse(item.matches.previous_module_equal)

    async def test_full_comparison_rule(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Chris",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        item = response.items[0]
        self.assertTrue(item.matches.count_equal)
        self.assertTrue(item.matches.any_equal)

    async def test_previous_module_true(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Alice",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertTrue(response.items[0].attendance.attended_previous_module)
        self.assertTrue(response.items[0].matches.previous_module_equal)

    async def test_previous_module_false(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Bob",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertFalse(response.items[0].attendance.attended_previous_module)
        self.assertFalse(response.items[0].matches.previous_module_equal)

    async def test_no_training_was_offered_maps_to_none(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search="Chris",
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
        )
        self.assertIsNone(response.items[0].matches.previous_module_equal)

    async def test_only_mismatches(self):
        response = await self.service.list_attendance_cross_check(
            project_id=self.project_id,
            page=1,
            page_size=10,
            search=None,
            training_group_id=None,
            verification_source="all",
            only_mismatches=True,
        )
        self.assertEqual(len(response.items), 1)
        self.assertFalse(response.items[0].is_match)

    async def test_export_all(self):
        content = await self.service.export_attendance_cross_check(
            project_id=self.project_id,
            search=None,
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
            export_scope="all",
        )
        wb = load_workbook(filename=io.BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.max_row, 4)

    async def test_export_mismatches(self):
        content = await self.service.export_attendance_cross_check(
            project_id=self.project_id,
            search=None,
            training_group_id=None,
            verification_source="all",
            only_mismatches=False,
            export_scope="mismatches",
        )
        wb = load_workbook(filename=io.BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.max_row, 2)


if __name__ == "__main__":
    unittest.main()
