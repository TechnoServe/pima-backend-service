from __future__ import annotations

import io
import unittest
from uuid import uuid4

from openpyxl import load_workbook

from app.domains.farmers.service import FarmersService


class FakeFarmersExportRepo:
    def __init__(self, country=None):
        self.country = country
        self.include_zimbabwe_farm_visit_data = None

    async def project_location_name(self, project_id):
        return self.country

    async def export_training_modules(self, project_id):
        return []

    async def export_farmers_base_rows(self, project_id, *, include_zimbabwe_farm_visit_data=False):
        self.include_zimbabwe_farm_visit_data = include_zimbabwe_farm_visit_data
        return [
            {
                "Project": "Project A",
                "first_name": "Ada",
                "middle_name": None,
                "last_name": "Lovelace",
                "gender": "Female",
                "age": 36,
                "number_of_trees": 100,
                "number_of_coffee_plots": 2,
                "farm_size": 1.5,
                "phone_number": "123",
                "other_id": "OID-1",
                "location": "Location A",
                "location_gps_latitude": None,
                "location_gps_longitude": None,
                "location_gps_altitude": None,
                "farmer_sf_id": "SF-1",
                "from_sf": True,
                "tns_id": "TNS-1",
                "hh_number": 1,
                "sf_household_id": "HH-1",
                "farmer_number": 1,
                "ffg_id": "FFG-1",
                "training_group": "Group A",
                "consent_provided": True,
                "fv_coffee_tree_numbers": 125,
                "date_of_latest_farm_visit": "2026-06-15",
                "reason_for_change_in_number_of_trees": "New trees planted",
                "status": "Active",
                "farmer_status": "Active",
                "farmer_trainer": "Trainer A",
                "business_advisor": "Advisor A",
                "create_in_commcare": True,
            },
            {
                "Project": "Project A",
                "first_name": "Grace",
                "last_name": "Hopper",
                "farmer_sf_id": "SF-2",
                "consent_provided": False,
                "status": "Inactive",
            },
            {
                "Project": "Project A",
                "first_name": "Katherine",
                "last_name": "Johnson",
                "farmer_sf_id": "SF-3",
                "consent_provided": None,
                "status": "Active",
            },
        ]

    async def export_attendance_map(self, **kwargs):
        return {}


class FarmersExportServiceTests(unittest.IsolatedAsyncioTestCase):
    async def test_export_includes_consent_provided_before_status(self):
        service = FarmersService(db=None)
        service.repo = FakeFarmersExportRepo()

        data = await service.export_excel(project_id=uuid4())
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        consent_idx = headers.index("consent_provided")

        self.assertEqual(headers[consent_idx + 1], "status")
        self.assertEqual(ws.cell(row=2, column=consent_idx + 1).value, "Yes")
        self.assertEqual(ws.cell(row=3, column=consent_idx + 1).value, "No")
        self.assertIsNone(ws.cell(row=4, column=consent_idx + 1).value)

    async def test_zimbabwe_export_includes_latest_farm_visit_tree_fields(self):
        service = FarmersService(db=None)
        repo = FakeFarmersExportRepo(country="Zimbabwe")
        service.repo = repo

        data = await service.export_excel(project_id=uuid4())
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        tree_idx = headers.index("number_of_trees")
        self.assertEqual(
            headers[tree_idx + 1 : tree_idx + 4],
            [
                "fv_coffee_tree_numbers",
                "date_of_latest_farm_visit",
                "reason_for_change_in_number_of_trees",
            ],
        )
        self.assertEqual(ws.cell(row=2, column=tree_idx + 2).value, 125)
        self.assertEqual(ws.cell(row=2, column=tree_idx + 3).value, "2026-06-15")
        self.assertEqual(ws.cell(row=2, column=tree_idx + 4).value, "New trees planted")
        self.assertTrue(repo.include_zimbabwe_farm_visit_data)


if __name__ == "__main__":
    unittest.main()
