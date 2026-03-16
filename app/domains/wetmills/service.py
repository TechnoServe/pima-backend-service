from __future__ import annotations

import csv
import tempfile

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from .repository import WetmillsRepository
from .schemas import PaginatedWetmillsResponse, WetmillsFilterOptionsResponse


class WetmillsService:
    ALLOWED_SURVEYS = [
        "manager_needs_assessment",
        "cpqi",
        "employees",
        "financials",
        "infrastructure",
        "kpis",
        "wet_mill_training",
        "waste_water_management",
        "water_and_energy_use",
    ]

    def __init__(self, db: AsyncSession):
        self.repo = WetmillsRepository(db)

    async def list_wetmills(
        self,
        *,
        programme: str,
        country: str | None,
        search: str | None,
        exporting_status: str | None,
        mill_status: str | None,
        page: int,
        page_size: int,
    ) -> PaginatedWetmillsResponse:
        rows, total, has_ownership = await self.repo.list_wetmills(
            programme=programme,
            country=country,
            search=search,
            exporting_status=exporting_status,
            mill_status=mill_status,
            page=page,
            page_size=page_size,
        )

        items = []
        for row in rows:
            payload = dict(row)
            if not has_ownership:
                payload["ownership_type"] = None
            items.append(payload)

        return PaginatedWetmillsResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
        )

    async def filter_options(self, *, programme: str, country: str | None) -> WetmillsFilterOptionsResponse:
        return WetmillsFilterOptionsResponse(
            **(await self.repo.filter_options(programme=programme, country=country))
        )

    @staticmethod
    def _export_headers() -> list[str]:
        return [
            "Wetmill ID",
            "Wetmill Name",
            "Country",
            "Programme",
            "Ownership",
            "Exporting Status",
            "Mill Status",
            "Manager Name",
            "Manager Role",
            "Registered On",
            "Created At",
            "Updated At",
        ]

    @staticmethod
    def _export_row(row: dict, has_ownership: bool) -> list[str]:
        return [
            str(row.get("wet_mill_unique_id") or ""),
            str(row.get("name") or ""),
            str(row.get("country") or ""),
            str(row.get("programme") or ""),
            str(row.get("ownership_type") or "") if has_ownership else "",
            str(row.get("exporting_status") or ""),
            str(row.get("mill_status") or ""),
            str(row.get("manager_name") or ""),
            str(row.get("manager_role") or ""),
            row.get("registration_date").isoformat() if row.get("registration_date") else "",
            row.get("created_at").isoformat() if row.get("created_at") else "",
            row.get("updated_at").isoformat() if row.get("updated_at") else "",
        ]

    @staticmethod
    def _resolve_question_value(row: dict) -> str:
        value = row.get("value_text")
        if value is None:
            value = row.get("value_number")
        if value is None:
            value = row.get("value_boolean")
        if value is None:
            value = row.get("value_date")
        if value is None:
            value = row.get("value_gps")
        return "" if value is None else str(value)

    @staticmethod
    def _sheet_group_key(row: dict) -> tuple[str, str, str, str, str]:
        visit_date = row.get("visit_date")
        completed_date = row.get("completed_date")
        return (
            str(row.get("wetmill_name") or ""),
            visit_date.isoformat() if visit_date else "",
            str(row.get("submitted_by") or ""),
            completed_date.isoformat() if completed_date else "",
            str(row.get("general_feedback") or ""),
        )

    async def export_excel(
        self,
        *,
        programme: str,
        country: str | None,
        search: str | None,
        exporting_status: str | None,
        mill_status: str | None,
    ) -> str:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()

        wb = Workbook(write_only=True)

        main_sheet = wb.create_sheet(title="Wetmills")
        main_sheet.append(self._export_headers())

        has_ownership = await self.repo.has_ownership_column()
        async for row in self.repo.stream_wetmills_for_export(
            programme=programme,
            country=country,
            search=search,
            exporting_status=exporting_status,
            mill_status=mill_status,
        ):
            main_sheet.append(self._export_row(row, has_ownership))

        for survey_type in self.ALLOWED_SURVEYS:
            question_names = await self.repo.list_survey_question_names_for_export(
                programme=programme,
                country=country,
                search=search,
                exporting_status=exporting_status,
                mill_status=mill_status,
                survey_type=survey_type,
            )

            sheet = wb.create_sheet(title=survey_type[:31])
            headers = [
                "Wetmill Name",
                "Visit Date",
                "Submitted By",
                "Completed Date",
                "General Feedback",
                *question_names,
            ]
            sheet.append(headers)

            current_key: tuple[str, str, str, str, str] | None = None
            current_payload: dict[str, str] | None = None

            async for row in self.repo.stream_survey_data_for_export(
                programme=programme,
                country=country,
                search=search,
                exporting_status=exporting_status,
                mill_status=mill_status,
                survey_type=survey_type,
            ):
                row_key = self._sheet_group_key(row)

                if current_key != row_key:
                    if current_payload is not None:
                        sheet.append([current_payload.get(h, "") for h in headers])

                    current_key = row_key
                    current_payload = {
                        "Wetmill Name": row_key[0],
                        "Visit Date": row_key[1],
                        "Submitted By": row_key[2],
                        "Completed Date": row_key[3],
                        "General Feedback": row_key[4],
                    }
                    for question_name in question_names:
                        current_payload[question_name] = ""

                question_name = str(row.get("question_name") or "").strip()
                if question_name and current_payload is not None:
                    current_payload[question_name] = self._resolve_question_value(row)

            if current_payload is not None:
                sheet.append([current_payload.get(h, "") for h in headers])

        wb.save(tmp.name)
        return tmp.name

    async def export_csv(
        self,
        *,
        programme: str,
        country: str | None,
        search: str | None,
        exporting_status: str | None,
        mill_status: str | None,
    ) -> str:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8")
        try:
            writer = csv.writer(tmp)
            writer.writerow(self._export_headers())

            has_ownership = await self.repo.has_ownership_column()
            async for row in self.repo.stream_wetmills_for_export(
                programme=programme,
                country=country,
                search=search,
                exporting_status=exporting_status,
                mill_status=mill_status,
            ):
                writer.writerow(self._export_row(row, has_ownership))
        finally:
            tmp.close()

        return tmp.name