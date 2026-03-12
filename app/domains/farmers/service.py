from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Dict, List
from uuid import UUID

from openpyxl import load_workbook, Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.gcs import upload_bytes, signed_get_url, download_bytes
from app.core.config import settings
from app.db.reflection import get_table
from app.domains.farmers.repository import FarmersRepository
from app.domains.farmers.models import UploadRun, UploadRowError
from app.shared.api_errors import (
    ConflictError,
    ExternalServiceError,
    NotFoundError,
    ValidationError,
)

from app.domains.farmers.schemas import (
    PaginatedFarmersResponse,
    FarmerListItem,
    FarmersFilterOptions,
    FilterOption,
    FarmersSummaryResponse,
    UploadValidationResult,
    UploadValidationWarning,
    UploadJob,
    UploadHistoryResponse,
    FailedRow,
)

from .models import table


def before_update(payload: dict, entity: str) -> dict:
    data = dict(payload or {})
    if "send_to_commcare" in table().c:
        data["send_to_commcare"] = True
    if "send_to_commcare_status" in table().c:
        data["send_to_commcare_status"] = "Pending"
    return data


def T(name: str):
    return get_table(name)


def pad2(n: int) -> str:
    return str(int(n)).zfill(2)


def build_household_name(hh_number: int) -> str:
    # 1 -> "01", 9 -> "09", 10 -> "10"
    return pad2(hh_number)


def build_household_tns_id(ffg_id: str, hh_number: int) -> str:
    # household 1 -> GGG2201, household 11 -> GGG2211
    return f"{ffg_id}{pad2(hh_number)}"


def build_household_composite(ffg_id: str, hh_number: int) -> str:
    # IMPORTANT: composite uses padded household number
    return f"{ffg_id}-{pad2(hh_number)}"


def build_farmer_composite(ffg_id: str, hh_number: int, farmer_number: int) -> str:
    # IMPORTANT: composite uses padded household number
    return f"{ffg_id}-{pad2(hh_number)}-{farmer_number}"


def build_tns_id(ffg_id: str, hh_number: int, farmer_number: int) -> str:
    # farmer: primary in household 1 -> GGG22011, secondary -> GGG22012
    return f"{ffg_id}{pad2(hh_number)}{farmer_number}"


@dataclass(frozen=True)
class NormalizedRow:
    row_number: int
    raw: dict

    farmer_identifier: str
    from_sf: bool | None

    ffg_id: str
    hh_number: int | None
    farmer_number: int | None

    status: str
    is_active_row: bool

    farmer_id: UUID | None


@dataclass(frozen=True)
class HouseholdKey:
    ffg_id: str
    hh_number: int


class FarmersService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = FarmersRepository(db)

    # ---------------- List Farmers ----------------
    async def list_farmers(
        self,
        *,
        project_id: UUID,
        page: int,
        page_size: int,
        search: Optional[str],
        gender: Optional[str],
        location_id: Optional[UUID],
        farmer_group_id: Optional[UUID],
        farmer_trainer_id: Optional[UUID],
        business_advisor_id: Optional[UUID],
        has_pending_commcare: Optional[bool],
        sort_by: str,
        sort_order: str,
    ) -> PaginatedFarmersResponse:
        rows, total = await self.repo.list_farmers(
            project_id=project_id,
            page=page,
            page_size=page_size,
            search=search,
            gender=gender,
            location_id=location_id,
            farmer_group_id=farmer_group_id,
            farmer_trainer_id=farmer_trainer_id,
            business_advisor_id=business_advisor_id,
            has_pending_commcare=has_pending_commcare,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        items: list[FarmerListItem] = []
        for r in rows:
            f = r["farmer"]
            m = getattr(f, "_mapping", f)
            full_name = " ".join([x for x in [m.get("first_name"), m.get("middle_name"), m.get("last_name")] if x])

            items.append(
                FarmerListItem(
                    id=m.get("id"),
                    first_name=m.get("first_name"),
                    middle_name=m.get("middle_name"),
                    last_name=m.get("last_name"),
                    full_name=full_name,
                    gender=str(m.get("gender")) if m.get("gender") else None,
                    age=m.get("age"),
                    phone_number=m.get("phone_number"),
                    tns_id=m.get("tns_id"),
                    farmer_group_id=m.get("farmer_group_id"),
                    farmer_group_name=r["farmer_group_name"],
                    household_id=m.get("household_id"),
                    household_number=r["household_number"],
                    is_primary_household_member=bool(m.get("is_primary_household_member")),
                    farmer_trainer_id=r.get("farmer_trainer_id"),
                    farmer_trainer_name=r.get("farmer_trainer_name"),
                    business_advisor_id=r.get("business_advisor_id"),
                    business_advisor_name=r.get("business_advisor_name"),
                    location_id=r["location_id"],
                    location_name=r["location_name"],
                    send_to_commcare=bool(m.get("send_to_commcare")),
                    send_to_commcare_status=str(m.get("send_to_commcare_status")) if m.get("send_to_commcare_status") else None,
                    updated_at=m.get("updated_at"),
                    created_at=m.get("created_at"),
                )
            )

        total_pages = (total + page_size - 1) // page_size
        return PaginatedFarmersResponse(items=items, total=total, page=page, page_size=page_size, total_pages=total_pages)

    async def summary(self, *, project_id: UUID) -> FarmersSummaryResponse:
        s = await self.repo.summary(project_id=project_id)
        return FarmersSummaryResponse(total=1, pending_commcare=s["pending_commcare"])

    async def filter_options(self, *, project_id: UUID) -> FarmersFilterOptions:
        FarmerGroup = T("farmer_groups")
        Location = T("locations")

        genders = [FilterOption(value="Male", label="Male"), FilterOption(value="Female", label="Female")]

        q_loc = (
            select(Location.c.id, (Location.c.location_name if "location_name" in Location.c else Location.c.name))
            .select_from(FarmerGroup)
            .join(Location, FarmerGroup.c.location_id == Location.c.id)
            .where(FarmerGroup.c.project_id == project_id)
            .distinct()
            .order_by((Location.c.location_name if "location_name" in Location.c else Location.c.name))
        )
        loc_rows = (await self.db.execute(q_loc)).all()

        fg_name_col = FarmerGroup.c.ffg_name if "ffg_name" in FarmerGroup.c else (FarmerGroup.c.name if "name" in FarmerGroup.c else FarmerGroup.c.id)
        q_fg = select(FarmerGroup.c.id, fg_name_col).where(FarmerGroup.c.project_id == project_id).order_by(fg_name_col)
        fg_rows = (await self.db.execute(q_fg)).all()

        return FarmersFilterOptions(
            genders=genders,
            locations=[FilterOption(value=str(r[0]), label=str(r[1])) for r in loc_rows],
            farmer_groups=[FilterOption(value=str(r[0]), label=str(r[1])) for r in fg_rows],
            farmer_trainers=[],
            business_advisors=[],
        )

    # ---------------- Export XLSX ----------------
    async def export_excel(self, *, project_id: UUID) -> bytes:
        project_location_name = await self.repo.project_location_name(project_id)
        location_name = (project_location_name or "").strip().lower()

        hide_coffee_plots = location_name in {"zimbabwe", "ethiopia"}
        use_farm_size_alias = location_name == "ethiopia"

        modules = await self.repo.export_training_modules(project_id)
        base_rows = await self.repo.export_farmers_base_rows(project_id)

        base_headers = [
            "num",
            "Project",
            "first_name",
            "middle_name",
            "last_name",
            "gender",
            "age",
            "number_of_trees",
            "number_of_coffee_plots",
            "farm_size",
            "phone_number",
            "coop_membership_number",
            "location",
            "location_gps_latitude",
            "location_gps_longitude",
            "farmer_sf_id",
            "from_sf",
            "tns_id",
            "hh_number",
            "sf_household_id",
            "farmer_number",
            "ffg_id",
            "training_group",
            "status",
            "farmer_status",
            "farmer_trainer",
            "business_advisor",
            "create_in_commcare",
        ]

        if not hide_coffee_plots:
            base_headers.insert(9, "number_of_coffee_plots")

        module_headers: List[str] = []
        for m in modules:
            module_number = m.get("module_number") or ""
            module_name = m.get("module_name") or ""
            module_key = m.get("sf_id") or m.get("id")
            module_headers.append(f"{module_number}-{module_name}-{module_key}")

        farmer_sf_ids = [str(r.get("farmer_sf_id") or "").strip() for r in base_rows if r.get("farmer_sf_id")]
        farmer_ids = [uuid for r in base_rows for uuid in (r.get("farmer_id"),) if r.get("farmer_id")]

        att_map = await self.repo.export_attendance_map(
            project_id=project_id,
            farmer_sf_ids=farmer_sf_ids,
            farmer_ids=farmer_ids,
            training_modules=modules,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Farmers"
        ws.append(base_headers + module_headers)

        for i, r in enumerate(base_rows, start=1):
            row = [
                i,
                r.get("Project") or "",
                r.get("first_name") or "",
                r.get("middle_name") or "",
                r.get("last_name") or "",
                r.get("gender") or "",
                r.get("age") if r.get("age") is not None else "",
                r.get("number_of_trees") if r.get("number_of_trees") is not None else None,
                r.get("number_of_coffee_plots") if r.get("number_of_coffee_plots") is not None else None,
                r.get("farm_size") if r.get("farm_size") is not None else None,
                r.get("phone_number") if r.get("phone_number") is not None else "",
                r.get("coop_membership_number") if r.get("coop_membership_number") is not None else "",
                r.get("location") or "",
                r.get("location_gps_latitude") if r.get("location_gps_latitude") is not None else "",
                r.get("location_gps_longitude") if r.get("location_gps_longitude") is not None else "",
                str(r.get("farmer_sf_id") or r.get("farmer_id") or ""),
                bool(r.get("from_sf")),
                r.get("tns_id") or "",
                r.get("hh_number") if r.get("hh_number") is not None else "",
                str(r.get("sf_household_id") or r.get("household_id") or ""),
                r.get("farmer_number") if r.get("farmer_number") is not None else "",
                r.get("ffg_id") or "",
                r.get("training_group") or "",
                r.get("status") or "",
                r.get("farmer_status") or "",
                r.get("farmer_trainer") or "",
                r.get("business_advisor") or "",
                bool(r.get("create_in_commcare")) if r.get("create_in_commcare") is not None else "",
            ]

            if not hide_coffee_plots:
                row.insert(9, r.get("number_of_coffee_plots") if r.get("number_of_coffee_plots") is not None else None)

            module_vals = []
            sfid = str(r.get("farmer_sf_id") or r.get("farmer_id") or "").strip()
            for m in modules:
                key = str(m.get("sf_id") or m.get("id"))
                module_vals.append(att_map.get((sfid, key), ''))
            ws.append(row + module_vals)

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # ---------------- Upload validate + run ----------------
    def validate_upload(self, *, file_bytes: bytes) -> UploadValidationResult:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_set = {h.lower(): i for i, h in enumerate(headers)}

        required = ["farmer_sf_id", "first_name", "last_name", "ffg_id", "hh_number", "farmer_number"]
        errors: list[UploadValidationWarning] = []
        for key in required:
            if key not in header_set:
                errors.append(
                    UploadValidationWarning(
                        type="missing_column",
                        message=f'Required column "{key}" is missing',
                        column=key,
                        severity="error",
                    )
                )

        total_rows = ws.max_row - 1 if ws.max_row else 0
        preview = []
        for row in ws.iter_rows(min_row=2, max_row=min(11, ws.max_row), values_only=True):
            obj = {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
            preview.append(obj)

        return UploadValidationResult(
            is_valid=len(errors) == 0,
            total_rows=max(total_rows, 0),
            preview_rows=preview,
            warnings=[],
            errors=errors,
        )

    async def start_upload(
        self,
        *,
        project_id: UUID,
        file_name: str,
        content_type: str | None,
        file_bytes: bytes,
        uploaded_by_id: UUID | None,
    ) -> UploadJob:
        active = await self.repo.get_active_upload(project_id=project_id)
        if active:
            raise ConflictError("Cannot upload: another run is currently processing.")

        blocking_parent = await self.repo.get_blocking_validation_parent(project_id=project_id)
        if blocking_parent:
            raise ConflictError(
                "Cannot upload new file while a parent run has validation errors. Reupload against that run.",
                details={"blocking_upload_id": str(blocking_parent.id)},
            )

        run = await self._create_upload_run(
            project_id=project_id,
            file_name=file_name,
            content_type=content_type,
            file_bytes=file_bytes,
            uploaded_by_id=uploaded_by_id,
            parent_upload_id=None,
        )
        await self._validate_or_queue_run(run=run, file_bytes=file_bytes)
        return await self.get_upload_job(run.id)

    async def _create_upload_run(
        self,
        *,
        project_id: UUID,
        file_name: str,
        content_type: str | None,
        file_bytes: bytes,
        uploaded_by_id: UUID | None,
        parent_upload_id: UUID | None,
    ) -> UploadRun:
        try:
            gcs = upload_bytes(
                project_id=str(project_id),
                category="farmer-uploads",
                filename=f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{file_name}",
                content=file_bytes,
                content_type=content_type,
            )
        except Exception as exc:
            raise ExternalServiceError("Failed to upload file to storage", details={"reason": str(exc)}) from exc

        run = UploadRun(
            project_id=project_id,
            filename=file_name,
            content_type=content_type,
            file_size_bytes=len(file_bytes),
            gcs_bucket=gcs["bucket"],
            gcs_object_name=gcs["object_name"],
            gcs_uri=gcs["gcs_uri"],
            status="validating",
            progress=0,  # CHANGED: validation starts at 0 (validation is 0-50)
            total_rows=0,
            success_count=0,
            failed_count=0,
            remaining_count=0,
            uploaded_by_id=uploaded_by_id,
            parent_upload_id=parent_upload_id,
            uploaded_at=datetime.utcnow(),
        )
        return await self.repo.create_upload_run(run)

    async def _validate_or_queue_run(self, *, run: UploadRun, file_bytes: bytes) -> None:
        run.status = "validating"
        run.progress = 10
        await self.db.commit()

        errors = await self._collect_validation_errors(run=run, file_bytes=file_bytes)
        if errors:
            await self._mark_validation_errored(run=run, errors=errors)
            return

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        run.status = "validating"
        run.progress = 50
        run.total_rows = max((ws.max_row - 1 if ws.max_row else 0), 0)
        run.success_count = 0
        run.failed_count = 0
        run.remaining_count = run.total_rows
        run.meta = {**(run.meta or {}), "validated": True, "queued_for_processing": True}
        await self.db.commit()

    # ---------------- NODE-STYLE VALIDATION ----------------
    async def _collect_validation_errors(self, *, run: UploadRun, file_bytes: bytes) -> list[dict]:
        issues: list[dict] = []

        base_validation = self.validate_upload(file_bytes=file_bytes)
        for err in base_validation.errors:
            issues.append({"row_number": 1, "field": err.column or "header", "error_type": err.type, "message": err.message})
        if issues:
            return issues

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_idx = {h.strip().lower(): i for i, h in enumerate(headers)}

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        normalized, row_level_issues = await self._normalize_rows(run=run, headers=headers, header_idx=header_idx, rows=rows)
        issues.extend(row_level_issues)
        if issues:
            return issues

        issues.extend(await self._validate_households(normalized))
        issues.extend(self._validate_duplicate_farmers(normalized))
        return issues

    async def _normalize_rows(
        self,
        *,
        run: UploadRun,
        headers: list[str],
        header_idx: dict[str, int],
        rows: list,
    ) -> tuple[list[NormalizedRow], list[dict]]:
        issues: list[dict] = []
        normalized: list[NormalizedRow] = []

        def add(row_number: int, field: str, message: str, error_type: str = "validation_error"):
            issues.append({"row_number": row_number, "field": field, "error_type": error_type, "message": message})

        for i, row in enumerate(rows, start=2):
            raw = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}

            farmer_identifier = str(self._cell(row, header_idx, "farmer_sf_id") or "").strip()
            if not farmer_identifier:
                add(i, "farmer_sf_id", "Missing farmer identifier")
                continue

            from_sf_raw = self._cell(row, header_idx, "from_sf")
            from_sf = None if from_sf_raw in (None, "") else str(from_sf_raw).strip().lower() in ("1", "true", "yes")

            ffg_id = str(self._cell(row, header_idx, "ffg_id") or "").strip()
            if not ffg_id:
                add(i, "ffg_id", "Missing ffg_id")
                continue

            hh_number_raw = self._cell(row, header_idx, "hh_number")
            farmer_number_raw = self._cell(row, header_idx, "farmer_number")

            hh_number: int | None = None
            farmer_number: int | None = None

            try:
                hh_number = int(hh_number_raw) if hh_number_raw not in (None, "") else None
            except Exception:
                add(i, "hh_number", "hh_number must be an integer")
                continue

            try:
                farmer_number = int(farmer_number_raw) if farmer_number_raw not in (None, "") else None
            except Exception:
                add(i, "farmer_number", "farmer_number must be an integer")
                continue

            if hh_number is None:
                add(i, "hh_number", "Missing hh_number")
                continue
            if farmer_number is None:
                add(i, "farmer_number", "Missing farmer_number")
                continue
            if farmer_number not in (1, 2):
                add(i, "farmer_number", "farmer_number must be 1 or 2")
                continue

            status = str(self._cell(row, header_idx, "status") or "Active").strip()
            is_active_row = status.lower() == "active"

            farmer_data = await self.repo.resolve_farmer_for_project(
                project_id=run.project_id,
                identifier=farmer_identifier,
                from_sf=from_sf,
                active_only=False,
            )
            farmer_id = farmer_data[0] if farmer_data else None
            if not farmer_id:
                add(i, "farmer_sf_id", "Farmer not found for this project")
                continue

            if is_active_row and not await self.repo.is_farmer_active(farmer_id=farmer_id):
                is_active_row = False

            normalized.append(
                NormalizedRow(
                    row_number=i,
                    raw=raw,
                    farmer_identifier=farmer_identifier,
                    from_sf=from_sf,
                    ffg_id=ffg_id,
                    hh_number=hh_number,
                    farmer_number=farmer_number,
                    status=status,
                    is_active_row=is_active_row,
                    farmer_id=farmer_id,
                )
            )

        return normalized, issues

    async def _validate_households(self, normalized: list[NormalizedRow]) -> list[dict]:
        issues: list[dict] = []

        def add(row_number: int, field: str, message: str):
            issues.append({"row_number": row_number, "field": field, "error_type": "validation_error", "message": message})

        active_rows = [r for r in normalized if r.is_active_row and r.hh_number is not None]

        groups: dict[HouseholdKey, list[NormalizedRow]] = {}
        for r in active_rows:
            key = HouseholdKey(r.ffg_id, int(r.hh_number))
            groups.setdefault(key, []).append(r)

        for key, members in groups.items():

            hh_number = pad2(key.hh_number)
            if len(members) > 2:
                for m in members:
                    add(m.row_number, "hh_number", f"Household {key.ffg_id}{hh_number} has more than 2 active members")
                continue

            primary_count = sum(1 for m in members if m.farmer_number == 1)
            if primary_count > 1:
                for m in members:
                    add(m.row_number, "farmer_number", f"Household {key.ffg_id}{hh_number} has more than one active primary member")
                continue

            if len(members) == 2:
                has_primary = any(m.farmer_number == 1 for m in members)
                has_secondary = any(m.farmer_number == 2 for m in members)
                if not (has_primary and has_secondary):
                    for m in members:
                        add(m.row_number, "farmer_number", f"Household {key.ffg_id}{hh_number} must have one primary (1) and one secondary (2) when 2 members are active")
                    continue

        return issues

    def _validate_duplicate_farmers(self, normalized: list[NormalizedRow]) -> list[dict]:
        issues: list[dict] = []
        seen: set[UUID] = set()

        def add(row_number: int, message: str):
            issues.append({"row_number": row_number, "field": "farmer_sf_id", "error_type": "validation_error", "message": message})

        for r in normalized:
            if not r.is_active_row:
                continue
            if not r.farmer_id:
                continue
            if r.farmer_id in seen:
                add(r.row_number, "Farmer appears multiple times as Active in the same upload")
                continue
            seen.add(r.farmer_id)

        return issues

    async def _mark_validation_errored(self, *, run: UploadRun, errors: list[dict]) -> None:
        report_bytes = self._build_validation_error_report(errors)
        try:
            uploaded = upload_bytes(
                project_id=str(run.project_id),
                category="farmer-upload-validation-errors",
                filename=f"{run.id}-validation-errors.xlsx",
                content=report_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            run.error_gcs_object_name = uploaded["object_name"]
            run.error_gcs_uri = uploaded["gcs_uri"]
        except Exception:
            pass

        run.status = "validation_errored"
        run.progress = 100
        run.total_rows = max(run.total_rows, len(errors))
        run.failed_count = len(errors)
        run.remaining_count = 0
        run.completed_at = datetime.utcnow()
        run.meta = {**(run.meta or {}), "validation_errors_count": len(errors)}

        self.db.add_all(
            [
                UploadRowError(
                    upload_run_id=run.id,
                    row_number=e["row_number"],
                    error_type=e.get("error_type", "validation_error"),
                    error_message=f"{e.get('field', 'row')}: {e.get('message', 'Validation error')}",
                )
                for e in errors
            ]
        )
        await self.db.commit()

    def _build_validation_error_report(self, errors: list[dict]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Errors"
        ws.append(["row_number", "field", "error_type", "message"])
        for e in errors:
            ws.append([e.get("row_number"), e.get("field"), e.get("error_type", "validation_error"), e.get("message")])
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # ---------------- Upload processing ----------------
    async def process_upload_run(self, *, upload_run_id: UUID, file_bytes: bytes | None = None) -> None:
        run = await self.repo.get_upload_run(upload_run_id)
        if not run or run.status != "processing":
            return

        group_id_by_ffg: dict[str, UUID] = {}

        try:
            if file_bytes is None:
                if not run.gcs_object_name:
                    raise ValidationError("Uploaded file location is missing")
                file_bytes = download_bytes(run.gcs_object_name)

            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active

            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            header_idx = {h.strip().lower(): i for i, h in enumerate(headers)}

            module_cols: list[tuple[int, str]] = []
            for idx, header in enumerate(headers):
                parts = (header or "").strip().split("-")
                if len(parts) >= 3:
                    module_cols.append((idx, parts[-1].strip()))

            modules = await self.repo.export_training_modules(run.project_id)
            module_uuid_by_key: Dict[str, UUID] = {str(m.get("sf_id") or m["id"]): m["id"] for m in modules}

            ffg_ids_in_file: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                ffg = self._cell(row, header_idx, "ffg_id")
                if ffg not in (None, ""):
                    ffg_ids_in_file.add(str(ffg).strip())

            if ffg_ids_in_file:
                await self._prefetch_group_ids_for_ffg_ids(
                    project_id=run.project_id,
                    ffg_ids=sorted(ffg_ids_in_file),
                    out_map=group_id_by_ffg,
                )

            run.total_rows = run.total_rows or max((ws.max_row - 1 if ws.max_row else 0), 0)
            run.remaining_count = run.total_rows
            run.status = "processing"
            run.progress = max(run.progress, 50)
            await self.db.commit()

            row_errors: list[UploadRowError] = []
            success_count = 0
            failed_count = 0

            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    async with self.db.begin_nested():
                        await self._process_single_row(
                            run=run,
                            row_number=row_number,
                            row=row,
                            headers=headers,
                            header_idx=header_idx,
                            module_cols=module_cols,
                            module_uuid_by_key=module_uuid_by_key,
                            row_errors=row_errors,
                            group_id_by_ffg=group_id_by_ffg,
                        )
                    success_count += 1
                except Exception as exc:
                    failed_count += 1
                    row_errors.append(
                        UploadRowError(
                            upload_run_id=run.id,
                            row_number=row_number,
                            farmer_id=None,
                            tns_id=str(self._cell(row, header_idx, "tns_id") or ""),
                            error_type="row_error",
                            error_message=str(exc),
                            raw_row={headers[i]: row[i] for i in range(min(len(headers), len(row)))},
                        )
                    )

                processed = success_count + failed_count
                run.success_count = success_count
                run.failed_count = failed_count
                run.remaining_count = max(run.total_rows - processed, 0)
                # CHANGED: processing is 50% of the bar (50 -> 100)
                run.progress = min(99, 50 + int((processed / max(run.total_rows, 1)) * 50))
                if processed % 25 == 0:
                    await self.db.commit()

            if row_errors:
                await self.repo.bulk_add_row_errors(row_errors)

            run.success_count = success_count
            run.failed_count = failed_count
            run.remaining_count = 0
            run.progress = 100
            run.status = "failed" if failed_count > 0 else "completed"
            run.completed_at = datetime.utcnow()
            await self.db.commit()
        except Exception as exc:
            await self._fail_run(run, message=str(exc))

    async def _prefetch_group_ids_for_ffg_ids(self, *, project_id: UUID, ffg_ids: list[str], out_map: dict[str, UUID]) -> None:
        FarmerGroup = T("farmer_groups")

        candidate_cols = []
        for c in ("ffg_id", "tns_id", "group_tns_id"):
            if c in FarmerGroup.c:
                candidate_cols.append(FarmerGroup.c[c])

        if not candidate_cols:
            return

        key_col = candidate_cols[0]

        q = (
            select(FarmerGroup.c.id, key_col)
            .where(FarmerGroup.c.project_id == project_id)
            .where(key_col.in_(ffg_ids))
        )
        rows = (await self.db.execute(q)).all()
        for gid, key in rows:
            if key is None:
                continue
            out_map[str(key).strip()] = gid

    async def _process_single_row(
        self,
        *,
        run: UploadRun,
        row_number: int,
        row,
        headers: list[str],
        header_idx: dict[str, int],
        module_cols: list[tuple[int, str]],
        module_uuid_by_key: Dict[str, UUID],
        row_errors: list[UploadRowError],
        group_id_by_ffg: dict[str, UUID],
    ) -> None:
        farmer_identifier = str(self._cell(row, header_idx, "farmer_sf_id") or "").strip()
        if not farmer_identifier:
            raise ValidationError("Missing farmer identifier")

        from_sf_raw = self._cell(row, header_idx, "from_sf")
        from_sf: bool | None = None
        if from_sf_raw not in (None, ""):
            from_sf = str(from_sf_raw).strip().lower() in ("1", "true", "yes")

        ffg_id = str(self._cell(row, header_idx, "ffg_id") or "").strip()
        if not ffg_id:
            raise ValidationError("Missing ffg_id")

        hh_number_raw = self._cell(row, header_idx, "hh_number")
        farmer_number_raw = self._cell(row, header_idx, "farmer_number")

        if hh_number_raw in (None, ""):
            raise ValidationError("Missing hh_number")
        if farmer_number_raw in (None, ""):
            raise ValidationError("Missing farmer_number")

        try:
            hh_number = int(hh_number_raw)
            farmer_number = int(farmer_number_raw)
        except Exception as exc:
            raise ValidationError("hh_number and farmer_number must be integers") from exc

        if farmer_number not in (1, 2):
            raise ValidationError("farmer_number must be 1 or 2")

        status = str(self._cell(row, header_idx, "status") or "Active").strip()
        is_active_row = status.lower() == "active"

        farmer_data = await self.repo.resolve_farmer_for_project(
            project_id=run.project_id,
            identifier=farmer_identifier,
            from_sf=from_sf,
            active_only=False,
        )
        if not farmer_data:
            raise NotFoundError("Farmer not found for this project")
        farmer_id, _current_group_id = farmer_data

        if is_active_row and not await self.repo.is_farmer_active(farmer_id=farmer_id):
            is_active_row = False

        # Resolve group id using cache
        target_group_id = group_id_by_ffg.get(ffg_id)
        if not target_group_id:
            target_group_id = await self.repo.resolve_group_by_tns(project_id=run.project_id, ffg_id=ffg_id)
            if target_group_id:
                group_id_by_ffg[ffg_id] = target_group_id

        if not target_group_id:
            raise ValidationError(f"Unknown ffg_id: {ffg_id}")

        # Find household by (group, number)
        household_id = await self.repo.find_household_by_group_number(
            farmer_group_id=target_group_id,
            household_number=hh_number,
        )

        if not household_id:
            # Create household + set created_by_id/last_updated_by_id to logged-in user (run.uploaded_by_id)
            household_values = self._build_household_values(
                target_group_id=target_group_id,
                run=run,
                ffg_id=ffg_id,
                hh_number=hh_number,
                row=row,
                header_idx=header_idx,
            )
            household_id = await self.repo.create_household(values=household_values)
        else:
            # Update household farmer_group_id too (supports group moves)
            household_updates = self._build_household_updates(
                target_group_id=target_group_id,
                ffg_id=ffg_id,
                hh_number=hh_number,
                row=row,
                header_idx=header_idx,
                last_updated_by_id=run.uploaded_by_id,
            )
            if household_updates:
                await self.repo.update_household(household_id=household_id, values=household_updates)

        # Enforce constraints only for active rows
        if is_active_row:
            member_count = await self.repo.count_household_members(household_id=household_id, exclude_farmer_id=farmer_id)
            if member_count >= 2:
                raise ValidationError("A household cannot have more than 2 active members")

            if farmer_number == 1:
                existing_primary = await self.repo.count_primary_members(household_id=household_id, exclude_farmer_id=farmer_id)
                if existing_primary >= 1:
                    raise ValidationError("A household cannot have more than one active primary member")

        # Farmer updates (ensure farmer_group_id + household_id are correct for moves)
        farmer_updates = self._build_farmer_updates(row=row, header_idx=header_idx, from_sf=from_sf)

        farmer_updates["farmer_group_id"] = target_group_id
        if "household_id" in T("farmers").c:
            farmer_updates["household_id"] = household_id
        if "is_primary_household_member" in T("farmers").c:
            farmer_updates["is_primary_household_member"] = (farmer_number == 1)

        farmer_comp = build_farmer_composite(ffg_id, hh_number, farmer_number)
        self._set_if_present(
            farmer_updates,
            T("farmers"),
            ["composite_id", "farmer_composite_id", "participant_composite_id"],
            farmer_comp,
        )

        computed_farmer_tns = build_tns_id(ffg_id, hh_number, farmer_number)
        if "tns_id" in T("farmers").c:
            farmer_updates["tns_id"] = computed_farmer_tns

        if farmer_updates:
            if "updated_at" in T("farmers").c:
                farmer_updates["updated_at"] = datetime.utcnow()
            if "send_to_commcare" in T("farmers").c:
                farmer_updates["send_to_commcare"] = True
            if "send_to_commcare_status" in T("farmers").c:
                farmer_updates["send_to_commcare_status"] = "Pending"
            await self.repo.update_farmer(farmer_id=farmer_id, values=farmer_updates)

        # Attendance columns (unchanged)
        for index, module_key in module_cols:
            if index >= len(row):
                continue

            value = row[index]
            if value in (None, ""):
                continue

            module_id = module_uuid_by_key.get(module_key)
            if not module_id:
                continue

            session_id = await self.repo.latest_session_id_for_group_module(
                farmer_group_id=target_group_id,
                module_id=module_id,
            )
            if not session_id:
                row_errors.append(
                    UploadRowError(
                        upload_run_id=run.id,
                        row_number=row_number,
                        farmer_id=farmer_id,
                        tns_id=str(self._cell(row, header_idx, "tns_id") or ""),
                        error_type="attendance_error",
                        error_message=f"No training session found for module {module_key}",
                        raw_row={headers[i]: row[i] for i in range(min(len(headers), len(row)))},
                    )
                )
                continue

            attended = str(value).strip().lower() in ("1", "true", "yes")
            attendance_values = (
                {"attended": attended}
                if "attended" in T("attendances").c
                else ({"status": "Present" if attended else "Absent"} if "status" in T("attendances").c else {})
            )

            attendance_id = await self.repo.attendance_id(
                project_id=run.project_id,
                farmer_id=farmer_id,
                session_id=session_id,
            )

            await self.repo.upsert_attendance(
                project_id=run.project_id,
                farmer_id=farmer_id,
                session_id=session_id,
                values=attendance_values,
                attendance_id=attendance_id,
                updated_by=run.uploaded_by_id,
            )

    def _build_farmer_updates(self, *, row, header_idx: dict[str, int], from_sf: bool | None) -> dict:
        farmer_table = T("farmers")
        updates: dict = {}

        editable_columns = [
            "first_name",
            "middle_name",
            "last_name",
            "gender",
            "age",
            "phone_number",
            "coop_membership_number",
            "status",
            "farmer_status",
            "create_in_commcare",
        ]

        for column in editable_columns:

            if column not in farmer_table.c:
                continue
            value = self._cell(row, header_idx, column)

            if column == "age":
                value = int(value)
            if column == "phone_number":
                value = str(value).strip()
            elif column == "create_in_commcare":
                value = str(value).strip().lower() in ("1", "true", "yes")
            updates[column] = value

        if "from_sf" in farmer_table.c and from_sf is not None:
            updates["from_sf"] = from_sf

        return updates

    def _build_household_values(
        self,
        *,
        target_group_id: UUID,
        run: UploadRun,
        ffg_id: str,
        hh_number: int,
        row,
        header_idx: dict[str, int],
    ) -> dict:
        household = T("households")
        values: dict = {}

        # Always set farmer_group_id on create (supports group moves)
        if "farmer_group_id" in household.c:
            values["farmer_group_id"] = target_group_id
        if "project_id" in household.c:
            values["project_id"] = run.project_id
        if "household_number" in household.c:
            values["household_number"] = hh_number

        # Audit fields on create (if columns exist)
        # Use the logged-in user id from the upload run.
        if run.uploaded_by_id:
            if "created_by_id" in household.c:
                values["created_by_id"] = run.uploaded_by_id
            if "last_updated_by_id" in household.c:
                values["last_updated_by_id"] = run.uploaded_by_id

        # Required household_name + household tns_id (recomputed deterministically)
        if "household_name" in household.c:
            values["household_name"] = build_household_name(hh_number)
        if "tns_id" in household.c:
            values["tns_id"] = build_household_tns_id(ffg_id, hh_number)

        # Composite IDs (padded household number)
        self._set_if_present(
            values,
            household,
            ["composite_id", "household_composite_id", "composite_household_id"],
            build_household_composite(ffg_id, hh_number),
        )

        values.update(self._household_shared_metrics(row=row, header_idx=header_idx))
        return values

    def _build_household_updates(
        self,
        *,
        target_group_id: UUID,
        ffg_id: str,
        hh_number: int,
        row,
        header_idx: dict[str, int],
        last_updated_by_id: UUID | None = None,
    ) -> dict:
        household = T("households")
        values: dict = {}

        # Always set farmer_group_id on update too (supports group moves)
        if "farmer_group_id" in household.c:
            values["farmer_group_id"] = target_group_id
        if "household_number" in household.c:
            values["household_number"] = hh_number

        # Keep audit field in sync on update (if present)
        if last_updated_by_id and "last_updated_by_id" in household.c:
            values["last_updated_by_id"] = last_updated_by_id

        if "household_name" in household.c:
            values["household_name"] = build_household_name(hh_number)
        if "tns_id" in household.c:
            values["tns_id"] = build_household_tns_id(ffg_id, hh_number)

        self._set_if_present(
            values,
            household,
            ["composite_id", "household_composite_id", "composite_household_id"],
            build_household_composite(ffg_id, hh_number),
        )

        values.update(self._household_shared_metrics(row=row, header_idx=header_idx))
        return values

    def _household_shared_metrics(self, *, row, header_idx: dict[str, int]) -> dict:
        household = T("households")
        values: dict = {}

        number_of_trees = self._cell(row, header_idx, "number_of_trees")
        if number_of_trees not in (None, "") and "number_of_trees" in household.c:
            if self._is_explicit_null_value(number_of_trees):
                values["number_of_trees"] = None
            else:
                values["number_of_trees"] = int(number_of_trees)

        coffee_plots = self._cell(row, header_idx, "number_of_coffee_plots")
        if coffee_plots not in (None, "") and "number_of_coffee_plots" in household.c:
            values["number_of_coffee_plots"] = None if self._is_explicit_null_value(coffee_plots) else coffee_plots

        farm_size = self._cell(row, header_idx, "farm_size")
        if farm_size not in (None, "") and "farm_size" in household.c:
            values["farm_size"] = None if self._is_explicit_null_value(farm_size) else farm_size

        if values and "updated_at" in household.c:
            values["updated_at"] = datetime.utcnow()

        return values

    @staticmethod
    def _set_if_present(target: dict, table_obj, candidate_columns: list[str], value):
        for col in candidate_columns:
            if col in table_obj.c:
                target[col] = value
                return

    @staticmethod
    def _cell(row, header_idx: dict[str, int], key: str):
        index = header_idx.get(key)
        if index is None or index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _is_explicit_null_value(value) -> bool:
        return isinstance(value, str) and value.strip().upper() == "NULL"

    async def _fail_run(self, run: UploadRun, *, message: str, failed_rows: int | None = None) -> None:
        run.status = "failed"
        run.completed_at = datetime.utcnow()
        run.progress = 100
        if failed_rows is not None:
            run.failed_count = failed_rows
            run.remaining_count = 0
            run.total_rows = max(run.total_rows, failed_rows)
        run.meta = {**(run.meta or {}), "error": message}
        await self.db.commit()

    # ---------------- Upload queries ----------------
    async def get_upload_job(self, upload_id: UUID) -> UploadJob:
        run = await self.repo.get_upload_run(upload_id)
        if not run:
            raise NotFoundError("Upload not found")

        latest = await self.repo.get_latest_upload_for_project(project_id=run.project_id)
        is_latest = bool(latest and latest.id == run.id)
        has_child = await self.repo.has_child_upload(upload_id=run.id)

        original_url = signed_get_url(run.gcs_object_name) if run.gcs_object_name else None
        error_url = signed_get_url(run.error_gcs_object_name) if run.error_gcs_object_name else None

        if not original_url and run.gcs_object_name:
            original_url = f"{settings.api_prefix}/farmers/uploads/{run.id}/original-file"
        if not error_url and run.error_gcs_object_name:
            error_url = f"{settings.api_prefix}/farmers/uploads/{run.id}/error-report"

        return UploadJob(
            id=run.id,
            project_id=run.project_id,
            filename=run.filename,
            status=run.status,
            progress=run.progress,
            total_rows=run.total_rows,
            success_count=run.success_count,
            failed_count=run.failed_count,
            remaining_count=run.remaining_count,
            uploaded_by_id=run.uploaded_by_id,
            uploaded_by_name=None,
            uploaded_at=run.uploaded_at,
            completed_at=run.completed_at,
            can_retry=is_latest and (run.status == "failed" or has_child),
            parent_upload_id=run.parent_upload_id,
            original_file_url=original_url,
            error_report_url=error_url,
        )

    async def active_upload(self, *, project_id: UUID) -> UploadJob | None:
        run = await self.repo.get_active_upload(project_id=project_id)
        if not run:
            return None
        return await self.get_upload_job(run.id)

    async def upload_history(self, *, project_id: UUID, page: int, page_size: int) -> UploadHistoryResponse:
        items, total = await self.repo.list_upload_history(project_id=project_id, page=page, page_size=page_size)
        jobs = [await self.get_upload_job(x.id) for x in items]
        total_pages = (total + page_size - 1) // page_size
        return UploadHistoryResponse(items=jobs, total=total, page=page, page_size=page_size, total_pages=total_pages)

    async def reupload_to_run(
        self,
        *,
        upload_id: UUID,
        file_name: str,
        content_type: str | None,
        file_bytes: bytes,
        uploaded_by_id: UUID | None,
    ) -> UploadJob:
        parent = await self.repo.get_upload_run(upload_id)
        if not parent:
            raise NotFoundError("Upload not found")

        if parent.status != "validation_errored":
            raise ConflictError("Reupload is only allowed for validation-errored runs")

        active = await self.repo.get_active_upload(project_id=parent.project_id)
        if active:
            raise ConflictError("Cannot reupload while another run is processing")

        child = await self._create_upload_run(
            project_id=parent.project_id,
            file_name=file_name,
            content_type=content_type,
            file_bytes=file_bytes,
            uploaded_by_id=uploaded_by_id,
            parent_upload_id=parent.id,
        )
        await self._validate_or_queue_run(run=child, file_bytes=file_bytes)
        return await self.get_upload_job(child.id)

    async def retry_upload(self, *, upload_id: UUID, mode: str) -> UploadJob:
        run = await self.repo.get_upload_run(upload_id)
        if not run:
            raise NotFoundError("Upload not found")
        if not run.gcs_object_name:
            raise ValidationError("Original upload file is missing")

        latest = await self.repo.get_latest_upload_for_project(project_id=run.project_id)
        is_latest = bool(latest and latest.id == run.id)
        has_child = await self.repo.has_child_upload(upload_id=run.id)
        if not is_latest:
            raise ConflictError("Only the most recent upload run in a project can be retried")
        if run.status != "failed" and not has_child:
            raise ConflictError("Retry is only allowed for failed runs or runs that have child reuploads")

        active = await self.repo.get_active_upload(project_id=run.project_id)
        if active:
            raise ConflictError("Cannot retry while another run is in progress")

        retry = UploadRun(
            project_id=run.project_id,
            filename=run.filename,
            content_type=run.content_type,
            file_size_bytes=run.file_size_bytes,
            gcs_bucket=run.gcs_bucket,
            gcs_object_name=run.gcs_object_name,
            gcs_uri=run.gcs_uri,
            status="validating",
            progress=0,
            total_rows=0,
            success_count=0,
            failed_count=0,
            remaining_count=0,
            uploaded_by_id=run.uploaded_by_id,
            parent_upload_id=run.id,
            uploaded_at=datetime.utcnow(),
            meta={"retry_mode": mode, "validated": True, "queued_for_processing": True, "retry_of": str(run.id)},
        )
        retry = await self.repo.create_upload_run(retry)
        return await self.get_upload_job(retry.id)

    async def queue_validated_runs_for_processing(self, *, limit: int = 5) -> int:
        q = (
            select(UploadRun)
            .where(
                UploadRun.status == "validating",
                UploadRun.meta["queued_for_processing"].astext == "true",
                UploadRun.meta["validated"].astext == "true",
            )
            .order_by(UploadRun.uploaded_at.asc())
            .limit(limit)
        )
        runs = list((await self.db.execute(q)).scalars().all())

        for run in runs:
            run.status = "processing"
            run.progress = 50

        if runs:
            await self.db.commit()
        return len(runs)

    async def get_upload_original_file_bytes(self, upload_id: UUID) -> tuple[bytes, str, str]:
        run = await self.repo.get_upload_run(upload_id)
        if not run:
            raise NotFoundError("Upload not found")
        if not run.gcs_object_name:
            raise ValidationError("Original upload file is missing")

        content = download_bytes(run.gcs_object_name)
        content_type = run.content_type or "application/octet-stream"
        filename = run.filename or f"upload-{upload_id}"
        return content, content_type, filename

    async def get_upload_error_report_bytes(self, upload_id: UUID) -> tuple[bytes, str, str]:
        run = await self.repo.get_upload_run(upload_id)
        if not run:
            raise NotFoundError("Upload not found")
        if not run.error_gcs_object_name:
            raise NotFoundError("No validation error report available for this run")

        content = download_bytes(run.error_gcs_object_name)
        return content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"validation-errors-{run.id}.xlsx"

    async def failed_rows(self, upload_id: UUID) -> list[FailedRow]:
        errs = await self.repo.list_failed_rows(upload_id)
        return [
            FailedRow(
                row_number=e.row_number,
                farmer_id=e.farmer_id,
                farmer_name=None,
                tns_id=e.tns_id,
                error_type=e.error_type,
                error_message=e.error_message,
            )
            for e in errs
        ]

    # ---------------- CommCare flagging ----------------
    async def send_to_commcare(self, *, project_id: UUID) -> int:
        return await self.repo.flag_farmers_send_to_commcare(project_id=project_id)

    async def pending_commcare_count(self, *, project_id: UUID) -> int:
        s = await self.repo.summary(project_id=project_id)
        return s["pending_commcare"]
